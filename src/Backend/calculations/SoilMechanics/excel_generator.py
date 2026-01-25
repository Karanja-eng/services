"""
Excel Export Module for Soil Mechanics
Professional calculation sheets with formulas and formatting
Uses openpyxl for full Excel support
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, ScatterChart, Reference, Series
from datetime import datetime
from typing import Dict, List, Optional
import io


class SoilMechanicsExcelExporter:
    """Generate professional Excel calculation sheets"""
    
    def __init__(self, project_info: Dict):
        self.project_info = project_info
        self.wb = Workbook()
        
        # Define styles
        self.header_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
        self.header_font = Font(bold=True, color="FFFFFF", size=12)
        self.subheader_fill = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")
        self.subheader_font = Font(bold=True, size=11)
        self.result_fill = PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid")
        self.result_font = Font(bold=True, size=11)
        
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        self.center_align = Alignment(horizontal='center', vertical='center')
        self.left_align = Alignment(horizontal='left', vertical='center')
    
    def generate_complete_workbook(
        self,
        phase_data: Optional[Dict] = None,
        atterberg_data: Optional[Dict] = None,
        compaction_data: Optional[Dict] = None,
        shear_data: Optional[Dict] = None,
        bearing_data: Optional[Dict] = None,
        consolidation_data: Optional[Dict] = None,
        slope_data: Optional[Dict] = None,
    ) -> bytes:
        """
        Generate complete Excel workbook with all test results
        Returns Excel file as bytes
        """
        # Remove default sheet
        if 'Sheet' in self.wb.sheetnames:
            del self.wb['Sheet']
        
        # Create summary sheet
        self._create_summary_sheet()
        
        # Create test sheets
        if phase_data:
            self._create_phase_sheet(phase_data)
        
        if atterberg_data:
            self._create_atterberg_sheet(atterberg_data)
        
        if compaction_data:
            self._create_compaction_sheet(compaction_data)
        
        if shear_data:
            self._create_shear_sheet(shear_data)
        
        if consolidation_data:
            self._create_consolidation_sheet(consolidation_data)
        
        if bearing_data:
            self._create_bearing_sheet(bearing_data)
        
        if slope_data:
            self._create_slope_sheet(slope_data)
        
        # Save to bytes
        buffer = io.BytesIO()
        self.wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()
    
    def _create_summary_sheet(self):
        """Create project summary sheet"""
        ws = self.wb.create_sheet("Project Summary", 0)
        
        # Title
        ws['A1'] = "GEOTECHNICAL INVESTIGATION"
        ws['A1'].font = Font(bold=True, size=16)
        ws['A1'].alignment = self.center_align
        ws.merge_cells('A1:F1')
        
        ws['A2'] = self.project_info.get('project_name', 'Project Name')
        ws['A2'].font = Font(bold=True, size=14)
        ws['A2'].alignment = self.center_align
        ws.merge_cells('A2:F2')
        
        # Project info table
        row = 4
        info_data = [
            ['Project Number:', self.project_info.get('project_no', 'N/A')],
            ['Client:', self.project_info.get('client', 'N/A')],
            ['Location:', self.project_info.get('location', 'N/A')],
            ['Engineer:', self.project_info.get('engineer', 'N/A')],
            ['Date:', datetime.now().strftime('%Y-%m-%d')],
        ]
        
        for label, value in info_data:
            ws[f'A{row}'] = label
            ws[f'A{row}'].font = self.subheader_font
            ws[f'B{row}'] = value
            row += 1
        
        # Tests performed
        row += 2
        ws[f'A{row}'] = "TESTS PERFORMED"
        ws[f'A{row}'].font = self.header_font
        ws[f'A{row}'].fill = self.header_fill
        ws.merge_cells(f'A{row}:D{row}')
        
        row += 1
        ws[f'A{row}'] = "Test Type"
        ws[f'B{row}'] = "Standard"
        ws[f'C{row}'] = "Status"
        ws[f'D{row}'] = "Sheet"
        
        for col in ['A', 'B', 'C', 'D']:
            ws[f'{col}{row}'].font = self.subheader_font
            ws[f'{col}{row}'].fill = self.subheader_fill
            ws[f'{col}{row}'].border = self.border
        
        # Column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
    
    def _create_phase_sheet(self, data: Dict):
        """Create phase relationships calculation sheet"""
        ws = self.wb.create_sheet("Phase Relationships")
        
        # Header
        ws['A1'] = "PHASE RELATIONSHIPS"
        ws['A1'].font = self.header_font
        ws['A1'].fill = self.header_fill
        ws.merge_cells('A1:D1')
        
        # Input section
        row = 3
        ws[f'A{row}'] = "INPUT PARAMETERS"
        ws[f'A{row}'].font = self.subheader_font
        ws[f'A{row}'].fill = self.subheader_fill
        ws.merge_cells(f'A{row}:C{row}')
        
        row += 1
        inputs = [
            ['Parameter', 'Value', 'Unit'],
            ['Mass of Wet Soil (Mw)', data.get('mass_wet', ''), 'g'],
            ['Mass of Dry Soil (Ms)', data.get('mass_dry', ''), 'g'],
            ['Specific Gravity (Gs)', data.get('specific_gravity', 2.65), '-'],
        ]
        
        for item in inputs:
            ws[f'A{row}'] = item[0]
            ws[f'B{row}'] = item[1]
            ws[f'C{row}'] = item[2]
            self._style_row(ws, row, ['A', 'B', 'C'])
            row += 1
        
        # Calculations
        row += 1
        ws[f'A{row}'] = "CALCULATIONS"
        ws[f'A{row}'].font = self.subheader_font
        ws[f'A{row}'].fill = self.subheader_fill
        ws.merge_cells(f'A{row}:D{row}')
        
        row += 1
        calcs = [
            ['Parameter', 'Formula', 'Value', 'Unit'],
            ['Moisture Content (w)', '=(B5-B6)/B6*100', data.get('moisture_content', ''), '%'],
            ['Void Ratio (e)', data.get('formula_e', ''), data.get('void_ratio', ''), '-'],
            ['Porosity (n)', '=e/(1+e)', data.get('porosity', ''), '-'],
            ['Bulk Unit Weight (γ)', '', data.get('gamma_bulk', ''), 'kN/m³'],
            ['Dry Unit Weight (γd)', '', data.get('gamma_dry', ''), 'kN/m³'],
            ['Saturated Unit Weight (γsat)', '', data.get('gamma_sat', ''), 'kN/m³'],
            ['Submerged Unit Weight (γ\')', '', data.get('gamma_sub', ''), 'kN/m³'],
        ]
        
        for item in calcs:
            ws[f'A{row}'] = item[0]
            ws[f'B{row}'] = item[1]
            ws[f'C{row}'] = item[2]
            ws[f'D{row}'] = item[3]
            
            if row > row - len(calcs) + 1:  # Results rows
                ws[f'C{row}'].fill = self.result_fill
                ws[f'C{row}'].font = self.result_font
            
            self._style_row(ws, row, ['A', 'B', 'C', 'D'])
            row += 1
        
        # Column widths
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 10
    
    def _create_atterberg_sheet(self, data: Dict):
        """Create Atterberg limits sheet"""
        ws = self.wb.create_sheet("Atterberg Limits")
        
        ws['A1'] = "ATTERBERG LIMITS (ASTM D4318)"
        ws['A1'].font = self.header_font
        ws['A1'].fill = self.header_fill
        ws.merge_cells('A1:D1')
        
        row = 3
        ws[f'A{row}'] = "TEST RESULTS"
        ws[f'A{row}'].font = self.subheader_font
        ws[f'A{row}'].fill = self.subheader_fill
        ws.merge_cells(f'A{row}:C{row}')
        
        row += 1
        results = [
            ['Parameter', 'Value', 'Unit'],
            ['Liquid Limit (LL)', data.get('liquid_limit', ''), '%'],
            ['Plastic Limit (PL)', data.get('plastic_limit', ''), '%'],
            ['Plasticity Index (PI)', '=B5-B6', '%'],
            ['USCS Classification', data.get('classification', ''), '-'],
            ['Description', data.get('description', ''), '-'],
        ]
        
        for item in results:
            ws[f'A{row}'] = item[0]
            ws[f'B{row}'] = item[1]
            ws[f'C{row}'] = item[2]
            
            if 'PI' in item[0] or 'Classification' in item[0]:
                ws[f'B{row}'].fill = self.result_fill
                ws[f'B{row}'].font = self.result_font
            
            self._style_row(ws, row, ['A', 'B', 'C'])
            row += 1
        
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 10
    
    def _create_compaction_sheet(self, data: Dict):
        """Create Proctor compaction sheet with chart"""
        ws = self.wb.create_sheet("Compaction")
        
        ws['A1'] = f"COMPACTION TEST - {data.get('test_type', 'Standard').upper()} PROCTOR"
        ws['A1'].font = self.header_font
        ws['A1'].fill = self.header_fill
        ws.merge_cells('A1:E1')
        
        # Test data table
        row = 3
        ws[f'A{row}'] = "TEST DATA"
        ws[f'A{row}'].font = self.subheader_font
        ws[f'A{row}'].fill = self.subheader_fill
        ws.merge_cells(f'A{row}:C{row}')
        
        row += 1
        ws[f'A{row}'] = "Moisture Content (%)"
        ws[f'B{row}'] = "Dry Density (kN/m³)"
        ws[f'C{row}'] = "Notes"
        
        for col in ['A', 'B', 'C']:
            ws[f'{col}{row}'].font = self.subheader_font
            ws[f'{col}{row}'].fill = self.subheader_fill
            ws[f'{col}{row}'].border = self.border
        
        row += 1
        start_data_row = row
        
        # Add test data points
        test_points = data.get('data_points', [])
        for point in test_points:
            ws[f'A{row}'] = point.get('moisture_content', '')
            ws[f'B{row}'] = point.get('dry_density', '')
            ws[f'C{row}'] = ''
            row += 1
        
        end_data_row = row - 1
        
        # Results
        row += 1
        ws[f'A{row}'] = "RESULTS"
        ws[f'A{row}'].font = self.subheader_font
        ws[f'A{row}'].fill = self.subheader_fill
        ws.merge_cells(f'A{row}:C{row}')
        
        row += 1
        ws[f'A{row}'] = "Maximum Dry Density (MDD)"
        ws[f'B{row}'] = data.get('maximum_dry_density', '')
        ws[f'C{row}'] = "kN/m³"
        ws[f'B{row}'].fill = self.result_fill
        ws[f'B{row}'].font = self.result_font
        
        row += 1
        ws[f'A{row}'] = "Optimum Moisture Content (OMC)"
        ws[f'B{row}'] = data.get('optimum_moisture_content', '')
        ws[f'C{row}'] = "%"
        ws[f'B{row}'].fill = self.result_fill
        ws[f'B{row}'].font = self.result_font
        
        # Create chart
        if len(test_points) > 0:
            chart = ScatterChart()
            chart.title = "Proctor Compaction Curve"
            chart.x_axis.title = "Moisture Content (%)"
            chart.y_axis.title = "Dry Density (kN/m³)"
            
            xvalues = Reference(ws, min_col=1, min_row=start_data_row, max_row=end_data_row)
            yvalues = Reference(ws, min_col=2, min_row=start_data_row, max_row=end_data_row)
            series = Series(yvalues, xvalues, title="Test Data")
            chart.series.append(series)
            
            ws.add_chart(chart, f"E{start_data_row}")
        
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 15
    
    def _create_shear_sheet(self, data: Dict):
        """Create shear strength sheet"""
        ws = self.wb.create_sheet("Shear Strength")
        
        ws['A1'] = "DIRECT SHEAR TEST (ASTM D3080)"
        ws['A1'].font = self.header_font
        ws['A1'].fill = self.header_fill
        ws.merge_cells('A1:D1')
        
        row = 3
        ws[f'A{row}'] = "TEST DATA"
        ws[f'A{row}'].font = self.subheader_font
        ws[f'A{row}'].fill = self.subheader_fill
        ws.merge_cells(f'A{row}:C{row}')
        
        row += 1
        ws[f'A{row}'] = "Normal Stress (kPa)"
        ws[f'B{row}'] = "Shear Stress (kPa)"
        
        for col in ['A', 'B']:
            ws[f'{col}{row}'].font = self.subheader_font
            ws[f'{col}{row}'].fill = self.subheader_fill
        
        row += 1
        
        # Results
        row += 5
        ws[f'A{row}'] = "SHEAR STRENGTH PARAMETERS"
        ws[f'A{row}'].font = self.subheader_font
        ws[f'A{row}'].fill = self.subheader_fill
        ws.merge_cells(f'A{row}:C{row}')
        
        row += 1
        results = [
            ['Cohesion (c)', data.get('cohesion', ''), 'kPa'],
            ['Friction Angle (φ)', data.get('friction_angle', ''), '°'],
            ['R² (Correlation)', data.get('r_squared', ''), '-'],
        ]
        
        for item in results:
            ws[f'A{row}'] = item[0]
            ws[f'B{row}'] = item[1]
            ws[f'C{row}'] = item[2]
            ws[f'B{row}'].fill = self.result_fill
            ws[f'B{row}'].font = self.result_font
            self._style_row(ws, row, ['A', 'B', 'C'])
            row += 1
        
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 10
    
    def _create_consolidation_sheet(self, data: Dict):
        """Create consolidation sheet"""
        ws = self.wb.create_sheet("Consolidation")
        
        ws['A1'] = "CONSOLIDATION TEST (ASTM D2435)"
        ws['A1'].font = self.header_font
        ws['A1'].fill = self.header_fill
        ws.merge_cells('A1:D1')
        
        row = 3
        ws[f'A{row}'] = "CONSOLIDATION PARAMETERS"
        ws[f'A{row}'].font = self.subheader_font
        ws[f'A{row}'].fill = self.subheader_fill
        ws.merge_cells(f'A{row}:C{row}')
        
        row += 1
        results = [
            ['Parameter', 'Value', 'Unit'],
            ['Compression Index (Cc)', data.get('compression_index', ''), '-'],
            ['Recompression Index (Cr)', data.get('recompression_index', ''), '-'],
            ['Preconsolidation Pressure (σ\'p)', data.get('preconsolidation_pressure', ''), 'kPa'],
            ['Primary Settlement (ΔH)', data.get('primary_settlement', ''), 'mm'],
        ]
        
        for item in results:
            ws[f'A{row}'] = item[0]
            ws[f'B{row}'] = item[1]
            ws[f'C{row}'] = item[2]
            
            if row > 4:
                ws[f'B{row}'].fill = self.result_fill
                ws[f'B{row}'].font = self.result_font
            
            self._style_row(ws, row, ['A', 'B', 'C'])
            row += 1
        
        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 10
    
    def _create_bearing_sheet(self, data: Dict):
        """Create bearing capacity sheet"""
        ws = self.wb.create_sheet("Bearing Capacity")
        
        ws['A1'] = f"BEARING CAPACITY - {data.get('method', 'TERZAGHI').upper()}"
        ws['A1'].font = self.header_font
        ws['A1'].fill = self.header_fill
        ws.merge_cells('A1:D1')
        
        row = 3
        ws[f'A{row}'] = "INPUT PARAMETERS"
        ws[f'A{row}'].font = self.subheader_font
        ws[f'A{row}'].fill = self.subheader_fill
        ws.merge_cells(f'A{row}:C{row}')
        
        row += 1
        inputs = [
            ['Parameter', 'Value', 'Unit'],
            ['Cohesion (c)', data.get('cohesion', ''), 'kPa'],
            ['Friction Angle (φ)', data.get('friction_angle', ''), '°'],
            ['Unit Weight (γ)', data.get('unit_weight', ''), 'kN/m³'],
            ['Footing Width (B)', data.get('width', ''), 'm'],
            ['Embedment Depth (Df)', data.get('depth', ''), 'm'],
        ]
        
        for item in inputs:
            ws[f'A{row}'] = item[0]
            ws[f'B{row}'] = item[1]
            ws[f'C{row}'] = item[2]
            self._style_row(ws, row, ['A', 'B', 'C'])
            row += 1
        
        row += 1
        ws[f'A{row}'] = "RESULTS"
        ws[f'A{row}'].font = self.subheader_font
        ws[f'A{row}'].fill = self.subheader_fill
        ws.merge_cells(f'A{row}:C{row}')
        
        row += 1
        results = [
            ['Ultimate Bearing Capacity (qu)', data.get('ultimate_bearing_capacity', ''), 'kPa'],
            ['Allowable Bearing Capacity (qa)', data.get('allowable_bearing_capacity', ''), 'kPa'],
            ['Factor of Safety', '3.0', '-'],
        ]
        
        for item in results:
            ws[f'A{row}'] = item[0]
            ws[f'B{row}'] = item[1]
            ws[f'C{row}'] = item[2]
            ws[f'B{row}'].fill = self.result_fill
            ws[f'B{row}'].font = self.result_font
            self._style_row(ws, row, ['A', 'B', 'C'])
            row += 1
        
        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 10
    
    def _create_slope_sheet(self, data: Dict):
        """Create slope stability sheet"""
        ws = self.wb.create_sheet("Slope Stability")
        
        ws['A1'] = f"SLOPE STABILITY - {data.get('method', 'BISHOP').upper()}"
        ws['A1'].font = self.header_font
        ws['A1'].fill = self.header_fill
        ws.merge_cells('A1:D1')
        
        row = 3
        ws[f'A{row}'] = "ANALYSIS RESULTS"
        ws[f'A{row}'].font = self.subheader_font
        ws[f'A{row}'].fill = self.subheader_fill
        ws.merge_cells(f'A{row}:C{row}')
        
        row += 1
        results = [
            ['Parameter', 'Value', 'Unit'],
            ['Factor of Safety (FoS)', data.get('factor_of_safety', ''), '-'],
            ['Status', data.get('status', ''), '-'],
            ['Analysis Method', data.get('method', 'Bishop'), '-'],
        ]
        
        for item in results:
            ws[f'A{row}'] = item[0]
            ws[f'B{row}'] = item[1]
            ws[f'C{row}'] = item[2]
            
            if 'FoS' in item[0]:
                ws[f'B{row}'].fill = self.result_fill
                ws[f'B{row}'].font = self.result_font
            
            self._style_row(ws, row, ['A', 'B', 'C'])
            row += 1
        
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 10
    
    def _style_row(self, ws, row, columns):
        """Apply standard styling to a row"""
        for col in columns:
            cell = ws[f'{col}{row}']
            cell.border = self.border
            cell.alignment = self.left_align


def export_to_excel(project_info: Dict, test_results: Dict) -> bytes:
    """
    Convenience function to export all results to Excel
    
    Args:
        project_info: Project metadata
        test_results: Dict containing all test results
    
    Returns:
        Excel file as bytes
    """
    exporter = SoilMechanicsExcelExporter(project_info)
    
    return exporter.generate_complete_workbook(
        phase_data=test_results.get('phase'),
        atterberg_data=test_results.get('atterberg'),
        compaction_data=test_results.get('compaction'),
        shear_data=test_results.get('shear'),
        bearing_data=test_results.get('bearing'),
        consolidation_data=test_results.get('consolidation'),
        slope_data=test_results.get('slope'),
    )